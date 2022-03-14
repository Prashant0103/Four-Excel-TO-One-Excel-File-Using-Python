from django.shortcuts import render,redirect
import openpyxl
from django.contrib import messages
import xlsxwriter
import pymysql
import os


def home(request):
    return render(request,'one.html')

def tryy(request):
    try:
        if request.method == 'POST':
            option = request.POST.get('module')
            fname = request.FILES['filename']
            con=pymysql.connect(host='localhost',user='root' , password='root',database='student')
            cur=con.cursor()
            book = openpyxl.load_workbook(fname)
            sheet_obj = book.active
            max_colu = sheet_obj.max_column
            max_roww = sheet_obj.max_row
            for i in range(2, max_roww + 1):
                A = []
                for j in range(2,max_colu+1):
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    A.append(cell_obj.value)
                
                if option=='student':
                        stmt1 = ("insert into stud_data(username,email,mobileno)" 
                        "values(%s,%s,%s)")
                        data1 = (A[0],A[1],A[2])
                        cur.execute(stmt1,data1)
                        con.commit()
                    
                elif option=='subject':
                    stmt2 = ("insert into subject(Name,Total_Marks)" 
                    "values(%s,%s)")
                    data2 = (A[0],A[1])
                    print(A)
                    cur.execute(stmt2,data2)
                    con.commit()
                    
                elif option=='stud_sub':
                    stmt3 = ("insert into student_subject(student_id,subject_id)" 
                    "values(%s,%s)")
                    data3 = (A[0],A[1])
                    print(A)
                    cur.execute(stmt3,data3)
                    con.commit()
                                    
                else:
                    print(option)
                    stmt4 = ("insert into student_report(Student_Subject_Id,Marks)" 
                    "values(%s,%s)")
                    data4 = (A[0],A[1])
                    print(A)
                    cur.execute(stmt4,data4)
                    con.commit()
            con.close()
            messages.success(request,'Data Stored Successfully..')
            return redirect('/')

    except:
        messages.error(request,'Data already exist or somthing went wrong')
        return redirect('/')
    
def export_ex(request):

    url = 'Student_report.xlsx'
    workbook = xlsxwriter.Workbook(url)
    worksheet = workbook.add_worksheet()
    print(worksheet.get_name())

    
    bold = workbook.add_format({'bold': True})

    worksheet.write('A1', 'ID', bold)
    worksheet.write('B1', 'UserName', bold)
    worksheet.write('C1', 'Email', bold)
    worksheet.write('D1', 'Mobile', bold)
    worksheet.write('E1', 'Php Marks', bold)
    worksheet.write('F1', 'Python Marks', bold)
    worksheet.write('G1', 'Cloud Computing Marks', bold)
    worksheet.write('H1', 'Total Marks', bold)
    worksheet.write('I1', 'OUT OF', bold)
    worksheet.write('J1', 'Percentage(%)', bold)


    conn=pymysql.connect(host='localhost',user='root' , password='root',database='student')
    cur=conn.cursor()
    cur1 = conn.cursor()

    sql = '''SELECT 
        stud_data.id,
        stud_data.username,
        stud_data.email,
        stud_data.mobileno,
        group_concat(student_report.Marks) as subjects_marks,
        SUM(student_report.Marks) as subject_total_marks,
        SUM(subject.Total_Marks) as total_marks
    FROM
        student_report
        LEFT JOIN
        student_subject ON student_subject.id = student_report.Student_Subject_Id
        LEFT JOIN
        subject ON subject.id = student_subject.Subject_Id
        LEFT JOIN
        stud_data ON stud_data.id = student_subject.Student_Id
        group by student_subject.student_id '''
    
    cur.execute(sql)

    A = cur.fetchall()
    for i in A:
        print(i)
        
    row = 1
    col = 0

    # Iterate over the data and write it out row by row.
    for id, uname,email,mobile,marks,ttmakrs,out_of in (A):
        worksheet.write(row, col,     id)
        worksheet.write(row, col + 1, uname)
        worksheet.write(row, col + 2, email)
        worksheet.write(row, col + 3, mobile)
        try:
            worksheet.write(row, col + 4, marks.split(',')[0])
        except:
            worksheet.write(row, col + 4, '--')
            continue
        try:
            worksheet.write(row, col + 5, marks.split(',')[1])
        except:
            worksheet.write(row, col + 5, '--')
            continue
        try:
            worksheet.write(row, col + 6, marks.split(',')[2])
        except:
            worksheet.write(row, col + 6, '--')
        worksheet.write(row, col + 7, ttmakrs)
        worksheet.write(row, col + 8, out_of)
        data = ttmakrs/out_of*100
        worksheet.write(row, col + 9, data)
        row += 1
    workbook.close()
    
    try:
        os.system('start "excel" "Student_report.xlsx"')
        messages.success(request,'Here We go')
        return redirect('/')
    
    except Exception as err:
        messages.error(request,err)
        return redirect('/')    
 

    
    

            
            